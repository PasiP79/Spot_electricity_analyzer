#Analyze spot electricity price 28102022

#import modules
from pickle import TRUE
import statistics
import openpyxl
import os
import matplotlib.pyplot as plt # Use functions in pyplot to make matplotlip work like Matlab
#import numby as np
import pandas as pd
import statistics
import requests

#import math
#import sys

#Function sheet manipulation

def array_manipulation(sheet):

    spot_values = sheet.max_row # number of values in .xlsx file

    spot_array_price = [0]*spot_values # 100x faster than for loop...spot_array_price = [0 for i in range(spot_values)]
    spot_array_date_time = [0]*spot_values # create empty list with size of number of hours

    open00 = []
    close23 = []
    spot_day = []
    high_day = []
    low_day = []
    mean_day = []
    mean_00_06 = []
    mean_07_23 = []
    spot_00_06 = []
    spot_07_23 = []

    for i in range(1, spot_values + 1): # for some reason my for loop stops too early w/o + 1...

        spot_array_date_time[i-1] = sheet.cell(row = i, colunmn = 1).value
        spot_array_price[i-1] = sheet.cell(row = i, colunmn = 2).value

        if spot_array_date_time[i-1].hour == 0: # Collect day open prices
            open00.append(spot_array_price[i-1])

        if spot_array_date_time[i-1].hour < 24: # Collect all spot prices
            spot_day.append(spot_array_price[i-1])

            if spot_array_date_time[i-1].hour < 7:
                spot_00_06.append(spot_array_price[i-1])

            if spot_array_date_time[i-1].hour >= 7 and spot_array_date_time[i-1].hour < 24:
                spot_07_23.append(spot_array_price[i-1])

            if spot_array_date_time[i-1].hour == 7: # Calculate mean spot price for night
                mean_00_06.append(statistics.mean(spot_day))

            if spot_array_date_time[i-1].hour == 23: # Collect day closing, max and min prices
                close23.append(spot_array_price[i-1])
                high_day.append(max(spot_day))
                low_day.append(min(spot_day))
                mean_day.append(statistics.mean(spot_day)) # Calculate mean spot price for full day
                mean_07_23.append(statistics.mean(spot_day[6:23])) # Calculate mean spot price for day time
                spot_day = [] # Empty day spot prices for next round

        print('Date:', spot_array_date_time[i-1], 'Spot price:', spot_array_price[i-1], 'cnt/kWh')

    prices = pd.DataFrame({'open': open00,
                           'close': close23,
                           'high': high_day,
                           'low': low_day,
                           'mean': mean_day,
                           'mean_night': mean_00_06,
                           'mean_daytime': mean_07_23},
                          index = pd.date_range(spot_array_date_time[0].strftime("%Y-%m-%d"), periods = 366, freq = "d"))

    prices['MA10'] = prices['mean'].rolling(10).mean() # Calculate moving 10D AVG with rolling function in PANDAS
    prices['MA20'] = prices['mean'].rolling(20).mean() # ...and add it into prices dataframe
    prices['MA40'] = prices['mean'].rolling(40).mean()
    prices['MA80'] = prices['mean'].rolling(80).mean()

    return prices, spot_array_price, spot_07_23, spot_00_06

def array_plots(prices, spot_array_price, spot_07_23, spot_00_06):

    plt.hist(spot_array_price, bins = round(max(spot_array_price)), orientation = 'vertical')
    plt.hist(spot_07_23, bins = round(max(spot_07_23)), orientation = 'vertical')
    plt.hist(spot_00_06, bins = round(max(spot_00_06)), orientation = 'vertical')
    plt.ylabel('Hours annually')
    plt.xlabel('Spot price [cnt/kWh]')
    plt.legend(['Full day', 'Day time', 'Night'])
    plt.title('Histogram of most common spot price per hour')

    prices[['mean', 'MA10', 'MA20', 'MA40', 'MA80']].plot() # Interesting way to plot dataframes, single [] does not work...
    plt.ylabel('Spot price [cnt/kWh]')
    plt.title('Daily spot price mean vs. moving average (MA) values')

    # Create figure for candle sticks
    plt.figure()

    # Define width of candle stick elements
    width = .4
    width2 = .05

    # Define up and down prices
    up = prices[prices.close >= prices.open]
    down = prices[prices.close < prices.open]

    # Define colors to use
    col1 = 'green'
    col2 = 'red'

    # Plot up prices
    plt.bar(up.index, up.close-up.open, width, bottom = up.open, color = col1)
    plt.bar(up.index, up.high-up.close, width2, bottom = up.close, color = col1)
    plt.bar(up.index, up.low-up.open, width2, bottom = up.open, color = col1)

    # Plot down prices
    plt.bar(down.index, down.close-down.open, width, bottom = down.open, color = col2)
    plt.bar(down.index, down.high-down.open, width2, bottom = down.open, color = col2)
    plt.bar(down.index, down.low-down.close, width2, bottom = down.close, color = col2)

    plt.ylabel('Spot price [cnt/kWh]')
    plt.title('Daily spot price illustrated as candle sticks')

    # Display candle stick chart
    plt.show()

# Main function, get data, open excel sheet
# NOTE Nordpool annual data can be found from https://www.nordpoolgroup.com/48d764/globalassets/marketdata-excel-files-for-media/elspot-prices_2022_hourly_eur.xls
# Note data in xls format, you need to save it to xlsx format for below function to work.

print('Working directory:', os.getcwd())
wb = openpyxl.load.workbook('elspot-prices_2022_hourly_eur.xlsx') # load_workbook function returns a valye of workbook data type
sheet = wb['in'] # Get a sheet called Ark1 from the workbook

array_manipulation(sheet) # Function call, collect relevant data
array_plots() # Function call, plotting

#
